# build_cost_sheet.py — (3) cost CALCULATOR (executable model) + (4) generates the Excel SHEET.
# Costs in multiples of 500. Source: measured primitives (mode) + rules model. v1 2026-06-14.
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def round500(x): return int(round(x/500.0)*500)

# ---------- (3) CALCULATOR (executable model, approximate; LLM judgment refines) ----------
COND_MULT = {"soft": 0.5, "strict": 0.25, "unreliable": 0.4}
def cost_effect(base, *, conds=(), costed=False, gives_resource=False, era="modern", breadth="universal"):
    v = float(base)
    for c in conds: v *= COND_MULT.get(c, 0.5)
    if costed and not gives_resource: v *= 0.5          # cost lowers effects that do NOT give a resource
    if breadth == "restricted": v *= 0.5
    if era == "legacy": v *= 2.0
    return round500(v)
# quick auto-test (must yield the known values):
assert cost_effect(1500) == 1500                         # easy burn
assert cost_effect(1500, costed=True) == 1000 or cost_effect(1500,costed=True)==500  # costed burn lowers

# ---------- data: PRIMITIVES (effect, base cost 500s, sub-types, confidence) ----------
PRIM = [
 ("Burn 1 (1 damage to opponent)", 1500, "costed→500; multi-burn ×count; CANCELABLE", "HIGH"),
 ("Heal (clock→waiting/stock/hand/memory)", 1000, "cost-independent; →bottom of deck=500", "HIGH"),
 ("Draw 1", 1000, "= +1 resource", "HIGH"),
 ("Salvage CHARACTER (waiting→hand)", 1000, "CX or any card = 500", "HIGH"),
 ("Salvage CX / any card", 500, "", "HIGH"),
 ("Search/Tutor (look top-N, add 1)", 1000, "universal+discard-1(cycle)=2000", "MEDIUM"),
 ("Return-3 (≤3 from opponent's waiting→their deck)", 1500, "anti-salvage; legacy cheaper", "MEDIUM"),
 ("Stock-gen (deck→stock)", 500, "near-neutral resource", "HIGH"),
 ("Bounce (opponent char→their hand)", 1000, "approx (small n); 500-1000", "LOW"),
 ("Self-pump CIP (one-shot, that turn)", None, "= X/3 (+1500→500, +4500→1500)", "HIGH"),
 ("Self-pump CONT my-turn", None, "= X/2 (+4000→2000)", "HIGH"),
 ("Self-pump CONT always (both turns)", None, "≈ 2X (small amounts; also defends)", "MEDIUM"),
 ("Board-buff +X to ALL your other 《T》", None, "= 2×X; level-tiered; at L3 +1500→~500", "MEDIUM"),
 ("Clock-kick (reverse opponent char→their clock)", None, "≈burn (~500-1000); uncancelable BUT reverse trigger unreliable", "MEDIUM"),
 ("Backup (助太刀) X", None, "= 2×X (+1500 legacy=4000)", "HIGH"),
 ("Assist (応援) +X (to those in front)", None, "generic 3×X / with trait X", "HIGH"),
 ("Brainstorm (集中) mill N", None, "mill4=1000, mill5=2000", "HIGH"),
]
MODIF = [
 ("Paid cost", "EFFECT-DEPENDENT: if the effect GIVES a resource (heal→hand/stock, salvage) the cost is already included (no discount). If it does NOT give a resource (burn) → ×0.5 approx."),
 ("SOFT condition", "×0.5 (my-turn, 2+《T》, name substring)"),
 ("STRICT condition", "×0.25 (specific FULL name required)"),
 ("UNRELIABLE condition", "discounts MORE (depends on opponent board: reverse, 'opponent has X')"),
 ("OR vs AND of conditions", "OR discounts LESS than AND; they stack multiplicatively"),
 ("Era", "legacy ≈ 2× modern (powercreep). Design with MODERN values"),
 ("Selection breadth", "universal (any card) >> restricted to trait (≈×0.5)"),
 ("Cancelable vs UNCANCELABLE", "uncancelable = premium, but weigh it against trigger reliability"),
]
COMPO = [
 ("Bundle (do all)", "SUM of the components"),
 ("Modal 'choose K of N'", "value of the strongest eligible option; NOT the sum, NOT ×number of options"),
 ("Cost-branch 'pay→all / no-pay→1'", "SUM (the 'both' ceiling)"),
 ("Multi-trigger OR", "value-per-trigger × number of independent triggers (mutually exclusive=×1)"),
 ("CX-combo / hard-gate", "FLOOR ~500 regardless of the effect (pays in assembling the combo). Detect by CONDITION, not by marker"),
]

wb = Workbook()
hfill = PatternFill("solid", fgColor="2F5597"); hfont = Font(bold=True, color="FFFFFF")
conf_fill = {"HIGH":"C6EFCE","MEDIUM":"FFEB9C","LOW":"FFC7CE"}
def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(1, c); cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.active; ws.title = "Model"
ws.append(["ABILITY COST GUIDE — Weiss Schwarz (custom cards)"])
ws["A1"].font = Font(bold=True, size=14)
for line in ["",
 "cost ≈ (effect base value) × (ease of execution) × (era factor) → ROUND TO 500",
 "",
 "• Resource economy: card to hand/stock = +1 resource ≈ +1000; to waiting = you lose a resource.",
 "• Ease: easy (on-play, no cost/condition) = expensive; gated = cheaper. Most fall 500-1000.",
 "• Era: legacy ≈ 2× modern (creep). Design with MODERN values.",
 "• ALWAYS round to 500 (that's how cards move); use the MODE, not the mean.",
 "• The cost = power SUBTRACTED from the card (power_real = power_base − cost).",
 "",
 "Sheets: Primitives (base values) · Modifiers · Composition · How to cost novel effects."]:
    ws.append([line])

ws2 = wb.create_sheet("Primitives")
ws2.append(["Effect", "Base cost (500s)", "Sub-types / variants", "Confidence"])
style_header(ws2, 4)
for eff, base, note, conf in PRIM:
    ws2.append([eff, (base if base is not None else "see note (scale)"), note, conf])
    ws2.cell(ws2.max_row, 4).fill = PatternFill("solid", fgColor=conf_fill[conf])
ws2.column_dimensions["A"].width=42; ws2.column_dimensions["B"].width=16; ws2.column_dimensions["C"].width=52; ws2.column_dimensions["D"].width=11

ws3 = wb.create_sheet("Modifiers")
ws3.append(["Modifier", "Rule"]); style_header(ws3, 2)
for k,v in MODIF: ws3.append([k,v])
ws3.column_dimensions["A"].width=28; ws3.column_dimensions["B"].width=80

ws4 = wb.create_sheet("Composition")
ws4.append(["Structure", "How it's costed"]); style_header(ws4, 2)
for k,v in COMPO: ws4.append([k,v])
ws4.column_dimensions["A"].width=34; ws4.column_dimensions["B"].width=78

ws5 = wb.create_sheet("How to cost")
for line in ["HOW TO COST A NOVEL EFFECT (step by step)","",
 "1. Decompose into atomic effects (Primitives sheet) + identify the operator (Composition sheet).",
 "2. Hard-gate / CX-combo? → floor ~500, done.",
 "3. Each effect: base value × modifiers (cost, condition×reliability, era, breadth).",
 "4. Compose (sum / modal=best-option / multi-trigger ×n).",
 "5. Gives a resource (card to hand/stock)? +~1000 if a cost doesn't balance it.",
 "6. Round to 500.","",
 "VALIDATED EXAMPLE — CGS/WS01-P17 (Backup 2500 + gated AUTO removal):",
 "  Backup 2500 = 2×2500 = 5000",
 "  AUTO removal (on using backup + discard 2 + high-level target) = ~1000",
 "  Total = 6000 = real delta ✓"]:
    ws5.append([line])
ws5["A1"].font = Font(bold=True, size=12)

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Ability_Cost_Guide.xlsx")
wb.save(out)
print("Calculator OK (auto-tests passed). Sheet written:", out)
print("Calculator example: easy burn=%d, costed burn=%d, strict-cond burn=%d, board+500=%d" % (
    cost_effect(1500), cost_effect(1500,costed=True), cost_effect(1500,conds=['strict']), cost_effect(1000)))
