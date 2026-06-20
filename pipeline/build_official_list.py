# build_official_list.py — FINAL DELIVERABLE in OFFICIAL format (6 sheets per TYPE, all in English).
# Costs: measured->residual->estimated (validated 98%). Confidence = METHOD. EN = official + translation.
# Example card = up to 3. BUDGET concept = power_base-500. Dumps to_translate.json + _tr/chunk_*.json.
import json, os, re, collections, statistics as st, glob, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import official_en

D = os.path.dirname(os.path.abspath(__file__))
def _nk(s):  # normalize encoding noise (full/half-width, quote styles, ALL spacing); does NOT change meaning
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", s or ""))
clean = json.load(open(os.path.join(D, "cardlist_clean.json"), encoding="utf-8"))
en_cards = json.load(open(os.path.join(D, "cardlist_en.json"), encoding="utf-8"))
era = json.load(open(os.path.join(D, "card_era.json"), encoding="utf-8"))
try:
    # PERMANENT translation cache, keyed by JP ability text (stable ground truth).
    # Keyed by JP -> survives any methodology/sig change. Translate once, reuse forever. NEVER delete this file.
    _rawcache = json.load(open(os.path.join(D, "translation_cache.json"), encoding="utf-8"))
except FileNotFoundError:
    _rawcache = {}
# look up by NORMALIZED jp text so encoding-noise variants (spacing/width/quotes) share one entry
CACHE = {_nk(k): v for k, v in _rawcache.items()}
print("translation cache loaded (permanent, JP-keyed):", len(_rawcache))

def pb(c):
    t = 1 if "soul" in (c.get("trigger") or []) else 0
    return 3000 + 2500*c["level"] + 1500*c["cost"] - 1000*t - 1000*(c["soul"]-1)
def ra(c):
    return [a for a in c["abilities"] if (a.get("text") or "").strip() not in ("-","ー","－","ｰ","")]
def base_num(cn):  # strip rarity/parallel suffix: DAL/W99-001SP -> DAL/W99-001 (same card, only art differs)
    return re.sub(r"(\d)[A-Za-z]+$", r"\1", cn or "")
ZT = str.maketrans("０１２３４５６７８９＋－", "0123456789+-")
TRAIT = re.compile(r"《[^》]*》"); NAME = re.compile(r"「[^」]*」")
def gen(t):
    t = t.translate(ZT); t = TRAIT.sub("《T》", t); t = NAME.sub("「N」", t)
    # trait COUNT does not affect cost (user rule): any trait restriction = 1 category. Collapse a list of
    # traits to a single 《T》; only NO-trait (generic) stays distinct (and pricier).
    t = re.sub(r"、?《T》(?:[かや・/／、]《T》)+", "《T》", t)
    return re.sub(r"\s+", " ", t).strip()
def r500(x): return int(round(x/500.0)*500)
def mode500(xs): return collections.Counter(r500(x) for x in xs).most_common(1)[0][0]

KW = {"助太刀":"Backup","応援":"Assist","集中":"Brainstorm","アンコール":"Encore","経験":"Experience",
      "記憶":"Memory","絆":"Bond","チェンジ":"Change","加速":"Accelerate","共鳴":"Resonance",
      "シフト":"Shift","大活躍":"Great Performance","フォース":"Force","ヒール":"Heal","バウンス":"Bounce"}
FAMPAT = [
  ("Burn", r"相手に\d+ダメージ"), ("Heal", r"自分のクロック[^。]{0,20}(控え室|ストック|手札|思い出)に置"),
  ("Clock Kick", r"相手のキャラ[^。]{0,20}(クロック置場|クロックに)置"),
  ("Bounce", r"相手のキャラ[^。]{0,12}手札に戻"), ("Return to Deck", r"相手の(控え室|キャラ)[^。]{0,20}山札に(戻|加え)"),
  ("Reverse Opp", r"相手のキャラ[^。]{0,12}【リバース】"), ("Opp Disrupt", r"相手の(手札|ストック|山札|思い出|レベル置場|クロック)"),
  ("Salvage", r"自分の(控え室|思い出)[^。]{0,22}手札に(戻す|加える)"), ("Search", r"山札[^。]{0,14}見[てる][^。]{0,28}(手札|加える)"),
  ("Look Deck", r"山札[^。]{0,4}(上から)?[^。]{0,6}\d+枚[^。]{0,8}見"), ("Comeback", r"(控え室|山札)[^。]{0,22}キャラ[^。]{0,14}舞台に置"),
  ("Stock Gen", r"(山札の上|デッキトップ|山札の上から)[^。]{0,12}ストック置場に置"), ("Draw", r"引く"),
  ("Add to Hand", r"手札に(加える|加え|戻す)"), ("Power Pump (board)", r"あなたの[^。]{0,16}キャラすべてに[^。]{0,8}パワーを[＋+]"),
  ("Power Pump (self)", r"このカードのパワーを[＋+]"), ("Power Pump", r"キャラ[^。]{0,10}パワーを[＋+]"),
  ("Power Debuff", r"パワーを[－\-]"), ("Soul", r"ソウルを[＋+\-－]"), ("Level", r"レベルを[＋+\-－]"),
  ("Grant Ability", r"』を与える|の能力を得"), ("Mill (self)", r"山札の上から\d+枚を[^。]{0,8}控え室"),
  ("Move", r"(前列|後列|別の枠|横の枠|の枠)に[^。]{0,6}(動かす|置く|移動)"), ("Stand/Rest", r"【スタンド】|【レスト】"),
  ("Stock Boost", r"ストック置場に置"), ("Choice", r"次の効果から|から\d+つを選"),
  ("Early Play", r"レベル\d+以下[^。]{0,12}手札からプレイ|レベルを参照しない"),
  ("Cannot Attack", r"アタックできない|サイドアタックできない"), ("Restriction", r"できない|選べない|受けない"),
  ("Card Select", r"\d+枚(まで)?選"),
]
# CX Combo: an ability HARD-GATED to a specific named climax. Detected on the gen()-normalized text
# (names already collapsed to 「N」) by the climax-area gate, NOT by the 【CXコンボ】 marker (legacy
# cards predate it). Two gate shapes + the explicit modern marker:
#   クライマックス置場に「N」が(ある|あり)  = "if [name] is in your climax area" (the classic combo trigger)
#   「N」が(クライマックス置場に)?置かれた     = "when [name] is placed (in the climax area)" (on-place flavor)
#   クライマックスコンボ / ＣＸコンボ / CXコンボ = the explicit tagged marker (kept for the few oddballs)
# Deliberately NOT matched: あなたのクライマックスが…置かれた ("when ANY/your climax is placed"), which is a
# generic on-climax trigger, not gated to a specific combo CX -> it must keep its own family.
CXC_PAT = re.compile(r"クライマックス置場に「N」が(ある|あり)|「N」が(クライマックス置場に)?置かれた|クライマックスコンボ|ＣＸコンボ|CXコンボ")
def family(text):
    if CXC_PAT.search(text): return "CX Combo"   # FIRST: a combo encapsulates whatever sub-effects it mixes
    for k, v in KW.items():
        if k in text: return v
    for name, pat in FAMPAT:
        if re.search(pat, text): return name
    return "Other"

def ability_type(markers):
    m = "".join(markers or "")
    if "永" in m: return "CONT"
    if "自" in m: return "AUTO"
    if "起" in m: return "ACT"
    return "OTHER"

# ---------- de-duplicate alt-art / rarity parallels ----------
# A physical card = same base number + name + stats + EXACT effects. Cards that share a number/name
# but differ in effect (warned by the user) keep their own row. Keep the plain base as representative.
def _card_key(c):
    return (base_num(c.get("card_number", "")), _nk(c.get("name")), c.get("power"), c.get("level"),
            c.get("cost"), c.get("soul"), tuple((a.get("type"), _nk(a.get("text"))) for a in ra(c)))
def _rep_better(cand, cur):
    b = base_num(cand.get("card_number", ""))
    cb, ub = cand.get("card_number") == b, cur.get("card_number") == b
    if cb != ub: return cb                                   # the plain base number wins
    return len(cand.get("card_number", "")) < len(cur.get("card_number", ""))  # else the shortest
_dedup = {}
for c in clean:
    k = _card_key(c)
    if k not in _dedup or _rep_better(c, _dedup[k]):
        _dedup[k] = c
_before = len(clean); clean = list(_dedup.values())
print(f"alt-art de-dup: {_before} rows -> {len(clean)} distinct cards (removed {_before - len(clean)})")

# ---------- collect ----------
EN_AB = official_en.build(clean, en_cards)
variant_occ = collections.defaultdict(list)   # sig -> [(card, idx, markers, text)]
iso = collections.defaultdict(lambda: {"m": [], "l": []})
variant_text = {}                              # sig -> (markers, gen_text)
cards = []
for c in clean:
    if c["type"] != "Character" or c["excluded"]: continue
    if c["power"] is None or c["level"] is None or c["cost"] is None or c["soul"] is None: continue
    ab = ra(c)
    if not ab: continue
    sigs = []
    for i, a in enumerate(ab):
        mk = "".join(a.get("markers") or [])
        sig = mk + " :: " + gen(a.get("text", ""))
        sigs.append(sig)
        variant_occ[sig].append((c["card_number"], i, mk, a.get("text", "")))
        variant_text.setdefault(sig, (mk, gen(a.get("text", ""))))
    delta = pb(c) - c["power"]
    e = era.get(c["card_number"])
    cards.append((c["card_number"], delta, sigs, e))
    if len(ab) == 1:
        (iso[sigs[0]]["m"] if e == "modern" else iso[sigs[0]]["l"]).append(delta)

ALLV = set(variant_occ)
CXC_FLOOR = 500   # a CX-combo ability is worth >= 500: the cost is paid by ASSEMBLING the combo, not in power
def is_cxc(sig): return family(variant_text[sig][1]) == "CX Combo"
# STEP 1 measured (a single-ability CX-combo card still measures directly)
cost = {}; method = {}; nsamp = {}; rng = {}
for sig, d in iso.items():
    use = d["m"] if len(d["m"]) >= 2 else (d["m"] + d["l"])
    if not use: continue
    # a single LEGACY-only isolated sample is era-inflated & unstable -> don't lock it as measured;
    # let the modern multi-card residual (or family estimate) give a sane value instead.
    if not d["m"] and len(d["l"]) == 1:
        continue
    cost[sig] = mode500(use); method[sig] = "measured"; nsamp[sig] = len(use); rng[sig] = (min(use), max(use))
# families that may legitimately cost NEGATIVE (drawbacks) = those seen negative in MEASURED data
neg_fams = {family(variant_text[s][1]) for s, c in cost.items() if c < 0}
# STEP 2 propagated residual -- NON-CXC only. Solve a card's lone unknown ONLY when it is NOT a
# CX-combo sig; a still-unknown CXC sig keeps the card "unresolved" here, deferring CXC so it becomes
# the residual ABSORBER later (CXC is the hardest to measure directly: gated + pay-to-assemble).
multi = [(cn, dl, sg, e) for (cn, dl, sg, e) in cards if len(sg) > 1]
for _ in range(10):
    res = collections.defaultdict(list)
    for cn, dl, sg, e in multi:
        unk = [s for s in sg if s not in cost]
        if len(unk) == 1 and not is_cxc(unk[0]):
            res[unk[0]].append(dl - sum(cost[s] for s in sg if s in cost))
    new = 0
    for sig, samples in res.items():
        if sig in cost: continue
        val = mode500(samples)
        # a beneficial ability can't be a drawback: a negative residual means the seeds over-counted
        # (the card is over-budget / a value card) -> reject it, let PASO 3 estimate a sane positive.
        if val < 0 and family(variant_text[sig][1]) not in neg_fams:
            continue
        cost[sig] = val; method[sig] = "residual"; nsamp[sig] = len(samples); rng[sig] = (min(samples), max(samples)); new += 1
    if new == 0: break
# validation (checkpoint: measured+residual only, estimated/unresolved excluded -- same basis as before)
errs = [abs(dl - sum(cost[s] for s in sg)) for cn, dl, sg, e in multi if all(s in cost for s in sg)]
if errs:
    print(f"VALIDATION: |err|<=500 in {sum(1 for x in errs if x<=500)/len(errs)*100:.0f}% (n={len(errs)})")
# STEP 3 estimated (family median). 3a: estimate every NON-CXC sig first, so the lone unknown left on a
# CX-combo card is its CXC sig. 3b: derive each CXC sig as the residual absorber (delta - sum(non-CXC)),
# floored. 3c: estimate any leftover (CXC -> floored CXC median, else family median).
fam_known = collections.defaultdict(list)
for sig, cst in cost.items():
    if not is_cxc(sig): fam_known[family(variant_text[sig][1])].append(cst)
fam_med = {f: r500(st.median(v)) for f, v in fam_known.items()}
for sig in ALLV:                                   # 3a non-CXC family estimate
    if sig in cost or is_cxc(sig): continue
    cost[sig] = fam_med.get(family(variant_text[sig][1]), 500); method[sig] = "estimated"; nsamp[sig] = 0; rng[sig] = (None, None)
for _ in range(10):                                # 3b CXC residual absorber
    res = collections.defaultdict(list)
    for cn, dl, sg, e in multi:
        unk = [s for s in sg if s not in cost]
        if len(unk) == 1 and is_cxc(unk[0]):
            res[unk[0]].append(dl - sum(cost[s] for s in sg if s in cost))
    new = 0
    for sig, samples in res.items():
        if sig in cost: continue
        cost[sig] = max(CXC_FLOOR, mode500(samples)); method[sig] = "residual"
        nsamp[sig] = len(samples); rng[sig] = (min(samples), max(samples)); new += 1
    if new == 0: break
cxc_known = [c for s, c in cost.items() if is_cxc(s)]   # 3c estimate leftovers
cxc_med = max(CXC_FLOOR, r500(st.median(cxc_known))) if cxc_known else CXC_FLOOR
fam_med["CX Combo"] = cxc_med
for sig in ALLV:
    if sig in cost: continue
    base = cxc_med if is_cxc(sig) else fam_med.get(family(variant_text[sig][1]), 500)
    cost[sig] = base; method[sig] = "estimated"; nsamp[sig] = 0; rng[sig] = (None, None)
for sig in ALLV:                                   # enforce the CXC floor on every CX-combo sig
    if is_cxc(sig) and cost[sig] < CXC_FLOOR: cost[sig] = CXC_FLOOR

CONF = {"measured": "HIGH", "residual": "MEDIUM", "estimated": "LOW"}

# debug:  DBG_CARD=RZ/SE35-11 python build_official_list.py   (per-ability breakdown of one card)
#         DBG_SIG="レベル×500" python build_official_list.py  (dump variants matching a substring + isolated samples)
if os.environ.get("DBG_CARD") or os.environ.get("DBG_SIG"):
    dbg = os.environ.get("DBG_CARD")
    if dbg:
        for c in clean:
            if c["card_number"] == dbg:
                print(f"=== {dbg}: pb={pb(c)} real={c['power']} delta={pb(c)-c['power']} era={era.get(dbg)} "
                      f"L{c['level']} C{c['cost']} S{c['soul']} trig={c.get('trigger')}")
                for i, a in enumerate(ra(c)):
                    mk = "".join(a.get("markers") or [])
                    sg = mk + " :: " + gen(a.get("text", ""))
                    print(f"  ab{i} [{ability_type(mk)}] cost={cost.get(sg)} method={method.get(sg)} "
                          f"ncards={len(variant_occ[sg])} | {(a.get('text') or '')[:100]}")
                break
    sub = os.environ.get("DBG_SIG")
    if sub:
        for sig in variant_text:
            if sub in sig:
                isos = [(cn, dl, e) for (cn, dl, sgs, e) in cards if len(sgs) == 1 and sgs[0] == sig]
                print(f"### cost={cost.get(sig)} method={method.get(sig)} fam={family(variant_text[sig][1])} "
                      f"ncards={len(variant_occ[sig])} iso={len(isos)}")
                print(f"    {sig[:140]}")
                print(f"    iso delta dist:", collections.Counter(dl for _, dl, _ in isos).most_common(8))
                print(f"    iso by era:", collections.Counter(e for _, _, e in isos).most_common())
                for cn, dl, e in sorted(isos)[:12]:
                    print(f"      {cn} delta={dl} era={e}")
    raise SystemExit

# ---------- per variant: type, examples, real JP, EN ----------
to_translate = []
rows_by_type = collections.defaultdict(list)
for sig in ALLV:
    occ = variant_occ[sig]
    off = next((o for o in occ if (o[0], o[1]) in EN_AB), None)
    if off:
        primary = off; jp_real = (primary[2] + " " + primary[3]).strip()
        en = EN_AB[(off[0], off[1])]                 # official EN exists -> free, no translation needed
    else:
        primary = occ[0]; jp_real = (primary[2] + " " + primary[3]).strip()
        en = CACHE.get(_nk(jp_real), "")             # normalized lookup; only blanks go to to_translate
    # examples up to 3 (primary first)
    seen = set(); ex = []
    for o in [primary] + occ:
        if o[0] not in seen:
            seen.add(o[0]); ex.append(o[0])
        if len(ex) >= 3: break
    typ = ability_type(primary[2])
    lo, hi = rng[sig]
    if not en:
        to_translate.append({"id": len(to_translate), "sig": sig, "type": typ, "jp": jp_real})
    rows_by_type[typ].append({
        "fam": family(variant_text[sig][1]), "en": en, "jp": jp_real, "power": cost[sig],
        "ncards": len(set(o[0] for o in occ)), "iso": (nsamp[sig] if method[sig] == "measured" else 0),
        "range": (f"{lo} – {hi}" if lo is not None else "—"), "conf": CONF[method[sig]],
        "method": method[sig], "ex": ", ".join(ex)})

# dump pending translations + chunks
json.dump(to_translate, open(os.path.join(D, "to_translate.json"), "w", encoding="utf-8"), ensure_ascii=False)
trdir = os.path.join(D, "_tr"); os.makedirs(trdir, exist_ok=True)
for f in glob.glob(os.path.join(trdir, "chunk_*.json")) + glob.glob(os.path.join(trdir, "out_*.json")):
    os.remove(f)   # transient per-run artifacts; the PERMANENT store is translation_cache.json (untouched)
CH = 120
nch = 0
for i in range(0, len(to_translate), CH):
    chunk = [{"id": r["id"], "type": r["type"], "jp": r["jp"]} for r in to_translate[i:i+CH]]
    json.dump(chunk, open(os.path.join(trdir, f"chunk_{nch:03d}.json"), "w", encoding="utf-8"), ensure_ascii=False)
    nch += 1
print(f"variants={len(ALLV)} | without EN (to translate)={len(to_translate)} | chunks={nch}")

# ---------- Excel formato OFICIAL ----------
ORDER = {"measured": 0, "residual": 1, "estimated": 2}
wb = Workbook()
HEAD = PatternFill("solid", fgColor="2F5597"); HF = Font(bold=True, color="FFFFFF")
SUBHEAD = PatternFill("solid", fgColor="D6DCE4")
CFILL = {"HIGH": "C6EFCE", "MEDIUM": "FFEB9C", "LOW": "D9D9D9"}
WRAP = Alignment(wrap_text=True, vertical="top")
CTOP = Alignment(vertical="top", horizontal="center")
TOP = Alignment(vertical="top")
HCEN = Alignment(wrap_text=True, vertical="top", horizontal="center")

def dlen(s):  # display width: CJK chars take ~2 cells
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in str(s))
def nlines(s, width):  # wrapped lines a string needs at a given column width
    return max(1, -(-dlen(s) // max(1, width - 2)) + str(s).count("\n"))
def fit_height(ws, row, pairs, cap=300, lh=15):  # pairs = [(text, colwidth), ...]
    ws.row_dimensions[row].height = min(lh * max(nlines(t, w) for t, w in pairs), cap)

def style_header(ws, n):
    for c in range(1, n+1):
        cell = ws.cell(1, c); cell.fill = HEAD; cell.font = HF; cell.alignment = HCEN

# Legend
ws = wb.active; ws.title = "Legend"
LEG = [
 ("Weiss Schwarz — Power cost per ability", ""),
 ("What this is", "How much POWER each Character ability costs (or gives, if a drawback), measured from the official Japanese card list. One row per distinct ability variant, split into sheets by ability type."),
 ("Power_base", "3000 + 2500*Level + 1500*Cost - 1000*[Trigger=Soul] - 1000*(Soul-1). Validated on 99.7% of vanilla characters."),
 ("Cost (delta)", "Power = Power_base - Power_real. POSITIVE = the ability costs power (it is beneficial). NEGATIVE = a drawback that gives power to the card. Always a multiple of 500."),
 ("BUDGET", "A card's BUDGET = Power_base - 500 (a card cannot exist at 0 power; 500 is the minimum). It is the total power a card can spend on abilities. Example: a Level 0 / Cost 0 / 1-soul / no soul-trigger card has Power_base 3000 -> budget 2500. To fit 3000 of abilities you must add a DRAWBACK ability that gives +500 power. See the 'Budget' sheet."),
 ("Sheets by type", "CONT = Continuous (永) · AUTO = Automatic (自) · ACT = Activated/Counter (起). OTHER = no standard type marker."),
 ("COLUMN", "MEANING"),
 ("Family", "Short label: the keyword (Backup, Assist, Brainstorm...) or a simple effect category (Burn, Heal, Salvage, Power pump...)."),
 ("Ability (English)", "Full English text of a representative card. Official Weiss Schwarz English where a verified match exists in the official EN card list; otherwise a faithful translation of the Japanese in official wording."),
 ("Source text (JP)", "The original full Japanese ability text of that representative card (the source of truth)."),
 ("Power", "Measured power cost of this ability (positive = costs power; negative = drawback that gives power)."),
 ("cards (n)", "How many distinct cards have this exact ability."),
 ("isolated (n)", "How many single-ability cards measured it DIRECTLY (power_base - real power). The basis of a HIGH-confidence value."),
 ("Range (min–max)", "Spread of the measured values across the sample (tight = consistent pricing; spread usually reflects different card budgets across generations, not error)."),
 ("Confidence", "Reflects the METHOD, not the number of samples. HIGH (green): MEASURED directly on single-ability card(s) — certain, even if only one card has the effect (more samples does NOT mean more confidence). MEDIUM (yellow): RESIDUAL — derived by subtracting known abilities from a multi-ability card. LOW (gray): ESTIMATED from the family median (no direct measurement)."),
 ("Example card", "Up to 3 representative cards with this ability (fewer if fewer exist). The shown text is from the first one."),
]
for a, b in LEG: ws.append([a, b])
ws["A1"].font = Font(bold=True, size=14)
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 95
for r in range(1, ws.max_row+1):
    ws.cell(r, 1).alignment = TOP
    ws.cell(r, 2).alignment = WRAP
    fit_height(ws, r, [(ws.cell(r, 2).value or "", 95)], cap=180)
# bold + shade the COLUMN/MEANING sub-header row
ws.cell(7, 1).font = Font(bold=True); ws.cell(7, 2).font = Font(bold=True)
ws.cell(7, 1).fill = SUBHEAD; ws.cell(7, 2).fill = SUBHEAD

# Budget sheet
wsb = wb.create_sheet("Budget")
wsb.append(["Vanilla power & Budget"]); wsb["A1"].font = Font(bold=True, size=14)
wsb.append([])
wsb.append(["VANILLA POWER (Power_base) is the printed power of a Character that has NO abilities. Every term in the formula below adds up to it — the soul-trigger and extra-soul terms are PART of the formula, not side notes."])
intro = wsb.max_row
wsb.merge_cells(start_row=intro, start_column=1, end_row=intro, end_column=4)
wsb.cell(intro, 1).alignment = WRAP
fit_height(wsb, intro, [(wsb.cell(intro, 1).value, 120)], cap=90)
wsb.append([])

# --- THE VANILLA POWER FORMULA, broken down term by term ---
wsb.append(["VANILLA POWER FORMULA"]); wsb.cell(wsb.max_row, 1).font = Font(bold=True, size=12)
wsb.append(["Term", "Value", "When it applies"])
hr = wsb.max_row
for c in range(1, 4):
    cell = wsb.cell(hr, c); cell.fill = HEAD; cell.font = HF; cell.alignment = HCEN
BRK = [
 ("Base", "3000", "every Character"),
 ("Level", "+ 2500 × Level", "per level printed on the card"),
 ("Cost", "+ 1500 × Cost", "per cost to play the card"),
 ("Soul trigger", "− 1000", "if the card's trigger icon is a Soul"),
 ("Extra Soul", "− 1000 × (Soul − 1)", "for each soul icon above the first"),
 ("Power_base (vanilla power)", "sum of all the terms above", "validated on 99.7% of vanilla (ability-less) characters"),
 ("BUDGET", "Power_base − 500", "the power a card can spend on abilities. A card cannot exist below 500 power, so 500 is held back. A DRAWBACK (negative-cost) ability raises the budget."),
]
for t, v, a in BRK:
    wsb.append([t, v, a]); rr = wsb.max_row
    wsb.cell(rr, 1).alignment = TOP; wsb.cell(rr, 2).alignment = TOP; wsb.cell(rr, 3).alignment = WRAP
    fit_height(wsb, rr, [(a, 60)], cap=60)
for rr in (wsb.max_row - 1, wsb.max_row):   # bold the two result rows
    wsb.cell(rr, 1).font = Font(bold=True); wsb.cell(rr, 2).font = Font(bold=True)
wsb.append([])

# --- worked table for the common case ---
wsb.append(["Power_base & Budget by stat line  (Soul 1, no soul-trigger):"]); wsb.cell(wsb.max_row, 1).font = Font(bold=True)
wsb.append(["Level", "Cost", "Power_base", "Budget (base − 500)"])
hr = wsb.max_row
for c in range(1, 5):
    cell = wsb.cell(hr, c); cell.fill = HEAD; cell.font = HF; cell.alignment = HCEN
for lv in range(0, 4):
    for co in range(0, 4):
        base = 3000 + 2500*lv + 1500*co
        wsb.append([lv, co, base, base-500])
        for c in range(1, 5): wsb.cell(wsb.max_row, c).alignment = Alignment(horizontal="center")
wsb.append([])
wsb.append(["Example:", "A Level 0 / Cost 0 card whose trigger is Soul and which has 2 soul = 3000 − 1000 (soul trigger) − 1000 (1 extra soul) = 1000 vanilla power  →  budget = 500."])
ex = wsb.max_row
wsb.cell(ex, 1).font = Font(bold=True); wsb.cell(ex, 1).alignment = TOP
wsb.merge_cells(start_row=ex, start_column=2, end_row=ex, end_column=4)
wsb.cell(ex, 2).alignment = WRAP
fit_height(wsb, ex, [(wsb.cell(ex, 2).value, 95)], cap=60)
for col, w in zip("ABCD", [30, 24, 60, 22]): wsb.column_dimensions[col].width = w

# sheets by type
for typ, title in [("CONT", "Power CONT"), ("AUTO", "Power AUTO"), ("ACT", "Power ACT"), ("OTHER", "Power OTHER")]:
    ws = wb.create_sheet(title)
    hdr = ["Family", "Ability (English)", "Source text (JP)", "Power", "cards (n)", "isolated (n)", "Range (min–max)", "Confidence", "Example card"]
    ws.append(hdr); style_header(ws, len(hdr))
    rows = sorted(rows_by_type.get(typ, []), key=lambda r: (r["fam"], ORDER[r["method"]], -r["power"]))
    for r in rows:
        ws.append([r["fam"], r["en"], r["jp"], r["power"], r["ncards"], r["iso"], r["range"], r["conf"], r["ex"]])
        rr = ws.max_row
        ws.cell(rr, 4).font = Font(bold=True)
        ws.cell(rr, 8).fill = PatternFill("solid", fgColor=CFILL[r["conf"]])
        for c in range(1, 10):
            ws.cell(rr, c).alignment = WRAP if c in (1, 2, 3, 7, 9) else CTOP
        # fix row height so the wrapped text shows vertically right away (don't rely on Excel auto-fit)
        fit_height(ws, rr, [(r["en"], 58), (r["jp"], 58), (r["ex"], 26)], cap=300)
    # widths kept moderate so wrapped text grows DOWN (vertical), not in one long horizontal line
    for col, w in zip("ABCDEFGHI", [18, 58, 58, 8, 8, 9, 13, 11, 26]): ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"   # filter/sort dropdowns on the headers

# Formulas (English, validated model)
wsf = wb.create_sheet("Formulas")
wsf.append(["Scalable cost formulas & design rules (validated from the data)"]); wsf["A1"].font = Font(bold=True, size=13)
wsf.append([])
wsf.append(["Family", "Template (X = the variable amount)", "Cost formula", "Notes / legacy exceptions"])
hr = wsf.max_row
for c in range(1, 5):
    cell = wsf.cell(hr, c); cell.fill = HEAD; cell.font = HF; cell.alignment = HCEN
FORMS = [
 ("Backup", "Act/Counter: BACKUP, your battling character gets +X power", "2 · X", "Legacy +1500 = 4000. To bottom of deck instead of waiting room: +1000 (a +3000 = 5000)."),
 ("Assist", "Cont: ASSIST, characters in front get +X (no restriction)", "3 · X", "Legacy +500 = 2000. ANY restriction (trait/level/name) -> X (1·X)."),
 ("Power pump (board)", "Cont: all your other <Trait> characters get +X", "≈ 3·X/2 + 500", "Tiered by level; collapses on L3."),
 ("Power pump (self, your turn)", "Cont: during your turn this card gets +X", "X / 2", "—"),
 ("Power pump (self, on-play one-shot)", "Auto: when placed, this card gets +X until end of turn", "X / 3", "—"),
 ("Power pump (self, always)", "Cont: this card gets +X (both turns)", "≈ 2·X", "Defends too -> more expensive (small amounts)."),
 ("Burn", "Auto: deal 1 damage to your opponent", "1500 easy / 1000 costed / 500 gated", "Cancelable by climax. Most are gated -> mode ≈ 500-1000."),
 ("Heal", "Auto: clock -> waiting/stock/hand/memory", "1000", "To bottom of deck = 500 (worse). Cost paid pays for the resource, doesn't discount."),
 ("Draw", "draw 1 card", "1000", "Net +1 card in hand."),
 ("Salvage", "choose 1 character in waiting room -> hand", "1000", "Any card / climax = 500."),
 ("Search / Tutor", "look at top N, add 1 to hand", "1000", "Universal + discard-cycle = 2000; restricted to trait ≈ half."),
 ("Brainstorm", "Act: BRAINSTORM, mill N", "mill4 = 1000, mill5 = 2000", "Salvage-brainstorm: any = 1000, trait = 500."),
 ("Stock gen", "deck -> stock", "500", "Near-neutral resource swap."),
 ("Bounce", "return an opponent's character to hand", "500 – 1000", "—"),
]
for f in FORMS:
    wsf.append(list(f)); rr = wsf.max_row
    for c in range(1, 5): wsf.cell(rr, c).alignment = WRAP
    fit_height(wsf, rr, [(f[0], 26), (f[1], 52), (f[2], 24), (f[3], 58)], cap=90)
wsf.append([])
DPRIN = [
 ("DESIGN PRINCIPLES", True),
 ("• Resource economy: a card to hand/stock ≈ +1 resource ≈ +1000; to waiting room = a lost resource. Cost PAYS for the resource.", False),
 ("• Era: legacy (<2017) ≈ 2× the modern cost (power creep). Design with MODERN values.", False),
 ("• Composition: bundle (do all) = SUM; modal (choose 1 of N) = the strongest option; multi-trigger = value × number of triggers.", False),
 ("• CX-combo / hard gate: floor ≈ 500 regardless of power (the cost is paid by assembling the combo).", False),
 ("• Conditions LOWER the cost (×1/2 soft, ×1/4 strict); OR-conditions discount less than AND.", False),
 ("• Sign of Power = benefit/drawback (negative = a drawback that gives power), not the family name.", False),
]
for text, bold in DPRIN:
    wsf.append([text]); rr = wsf.max_row
    wsf.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    wsf.cell(rr, 1).alignment = WRAP
    if bold: wsf.cell(rr, 1).font = Font(bold=True, size=12)
    fit_height(wsf, rr, [(text, 150)], cap=75)
for col, w in zip("ABCD", [26, 52, 24, 58]): wsf.column_dimensions[col].width = w

# SAFETY NET: never let a text cell be written as a formula (a leading "=" makes openpyxl
# emit <f>, which Excel then strips with a "removed records: formula" repair on open).
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                cell.data_type = "s"

out = os.path.join(D, "Complete_Abilities_List.xlsx"); wb.save(out)
nrows = sum(len(rows_by_type.get(t, [])) for t in ["CONT","AUTO","ACT","OTHER"])
withen = sum(1 for t in rows_by_type for r in rows_by_type[t] if r["en"])
print(f"rows={nrows} | with EN={withen} | without EN={nrows-withen}")
print("by type:", {t: len(rows_by_type.get(t, [])) for t in ["CONT","AUTO","ACT","OTHER"]})
print("escrito:", out)
