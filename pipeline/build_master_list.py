# build_master_list.py
# ============================================================================================
#  ⚠️  SUPERSEDED — DO NOT treat this file as the canonical cost source.
# ============================================================================================
# This is an OLDER, self-contained cost variant kept for reference/history. It predates and
# DIVERGES from the current model in pipeline/cost_model.py:
#   - it does NOT import cost_model.py (its cost math is a private copy that has since drifted),
#   - it has NO CX-combo family and NO replay-ability folding,
#   - its family labels + FAMPAT regexes differ from the canonical taxonomy,
#   - its era logic reads card_era.json expecting the old "legacy"/"modern" labels, but that file
#     now stores FORMAT names (Genesis/Bounty/…); so the `e == "modern"` test below is always
#     False and every isolated sample falls into the "legacy" bucket (harmless here only because
#     this file is not wired into the shipping flow — see cost_model.py for the live behaviour).
#   - it writes the SAME filename as build_official_list.py (Complete_Abilities_List.xlsx), so
#     running it will OVERWRITE the canonical Excel with this divergent version.
# The canonical deliverables are produced by build_official_list.py (Excel) and build_db.py
# (SQLite), both of which import cost_model.py. Regenerate those, not this. Kept only so the
# original measured->residual->estimated approach stays readable.
# ============================================================================================
#
# What it does (historical): MASTER LIST of ALL abilities (variants) with a cost per row.
# Method per row: MEASURED (single-ability cards, direct delta) -> RESIDUAL (subtract known seeds,
# with propagation) -> ESTIMATED (family median). Reliable official EN via official_en. All in 500s.
import json, os, re, csv, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import official_en

D = os.path.dirname(os.path.abspath(__file__))
clean = json.load(open(os.path.join(D, "cardlist_clean.json"), encoding="utf-8"))
en_cards = json.load(open(os.path.join(D, "cardlist_en.json"), encoding="utf-8"))
era = json.load(open(os.path.join(D, "card_era.json"), encoding="utf-8"))  # {card_number: legacy|modern}

# pb = vanilla ("base") power: what a Character with NO abilities would print at, per the game's
# power economy. delta (below) = pb - real power = the total power the card "spent" on its abilities.
def pb(c):
    t = 1 if "soul" in (c.get("trigger") or []) else 0
    return 3000 + 2500*c["level"] + 1500*c["cost"] - 1000*t - 1000*(c["soul"]-1)
def ra(c):   # real abilities: drop vanilla/placeholder dashes
    return [a for a in c["abilities"] if (a.get("text") or "").strip() not in ("-","ー","－","")]
ZT = str.maketrans("０１２３４５６７８９＋－", "0123456789+-")   # full-width -> ASCII
TRAIT = re.compile(r"《[^》]*》"); NAME = re.compile(r"「[^」]*」")
def gen(t):
    # Generalize an ability into a SIGNATURE: replace the card-specific trait 《…》 and name 「…」 with
    # placeholders so "gets +1000 for each 《Magic》" and "…each 《Weapon》" collapse to one variant.
    t = t.translate(ZT); t = TRAIT.sub("《T》", t); t = NAME.sub("「N」", t)
    return re.sub(r"\s+", " ", t).strip()
def r500(x): return int(round(x/500.0)*500)                       # snap to the nearest 500 (the game's granularity)
def mode500(xs): return collections.Counter(r500(x) for x in xs).most_common(1)[0][0]   # most common 500-rounded value

# ---------- FAMILY (transparent regex, broadened to reduce "Other") ----------
KW = {"助太刀":"Backup","応援":"Assist","集中":"Brainstorm","アンコール":"Encore","経験":"Experience",
      "記憶":"Memory","絆":"Bond","チェンジ":"Change","加速":"Accelerate","共鳴":"Resonance",
      "シフト":"Shift","大活躍":"GreatPerformance","フォース":"Force","ヒール":"Heal","バウンス":"Bounce"}
FAMPAT = [
  ("Burn", r"相手に\d+ダメージ"),
  ("Heal", r"自分のクロック[^。]{0,20}(控え室|ストック|手札|思い出)に置"),
  ("ClockKick", r"相手のキャラ[^。]{0,20}(クロック置場|クロックに)置"),
  ("Bounce", r"相手のキャラ[^。]{0,12}手札に戻"),
  ("ReturnDeck", r"相手の(控え室|キャラ)[^。]{0,20}山札に(戻|加え)"),
  ("ReverseOpp", r"相手のキャラ[^。]{0,12}【リバース】"),
  ("OppDisrupt", r"相手の(手札|ストック|山札|思い出|レベル置場|クロック)"),
  ("Salvage", r"自分の(控え室|思い出)[^。]{0,22}手札に(戻す|加える)"),
  ("Search", r"山札[^。]{0,14}見[てる][^。]{0,28}(手札|加える)"),
  ("LookDeck", r"山札[^。]{0,4}(上から)?[^。]{0,6}\d+枚[^。]{0,8}見"),
  ("Comeback", r"(控え室|山札)[^。]{0,22}キャラ[^。]{0,14}舞台に置"),
  ("StockGen", r"(山札の上|デッキトップ|山札の上から)[^。]{0,12}ストック置場に置"),
  ("Draw", r"引く"),
  ("AddToHand", r"手札に(加える|加え|戻す)"),
  ("PowerBoardAll", r"あなたの[^。]{0,16}キャラすべてに[^。]{0,8}パワーを[＋+]"),
  ("PowerSelf", r"このカードのパワーを[＋+]"),
  ("PowerOther", r"キャラ[^。]{0,10}パワーを[＋+]"),
  ("PowerDebuff", r"パワーを[－\-]"),
  ("SoulMod", r"ソウルを[＋+\-－]"),
  ("LevelMod", r"レベルを[＋+\-－]"),
  ("GrantAbility", r"』を与える|の能力を得"),
  ("MillSelf", r"山札の上から\d+枚を[^。]{0,8}控え室"),
  ("Move", r"(前列|後列|別の枠|横の枠|の枠)に[^。]{0,6}(動かす|置く|移動)"),
  ("StandRest", r"【スタンド】|【レスト】"),
  ("StockBoost", r"ストック置場に置"),
  ("Choice", r"次の効果から|から\d+つを選"),
  ("EarlyPlay", r"レベル\d+以下[^。]{0,12}手札からプレイ|レベルを参照しない"),
  ("CannotAttack", r"アタックできない|サイドアタックできない"),
  ("CannotEffect", r"できない|選べない|受けない"),
  ("CardSelect", r"\d+枚(まで)?選"),
]
def family(text):
    # Classify an ability into a family. Keyword hits (助太刀=Backup, …) win first because they are
    # unambiguous; otherwise fall through the ordered FAMPAT regexes (order matters — the first match
    # wins, so more specific patterns are listed above broader ones like "引く"=Draw). No match -> "Other".
    for k, v in KW.items():
        if k in text: return v
    for name, pat in FAMPAT:
        if re.search(pat, text): return name
    return "Other"

# ---------- collect valid cards: delta + sigs of each ability ----------
EN_AB = official_en.build(clean, en_cards)       # {(jp_code, ability_idx): en_text} reliable
cards = []                                       # (card_number, delta, [sig...], era)
variant_cards = collections.defaultdict(list)    # sig -> [card_number...]  (which cards carry this ability)
variant_occ = collections.defaultdict(list)      # sig -> [(card_number, idx)...]  (where, incl. position)
iso = collections.defaultdict(lambda: {"m": [], "l": []})  # sig -> isolated deltas, split modern/legacy
variant_text = {}                                # sig -> (markers, gen_text)  (representative display text)
for c in clean:
    if c["type"] != "Character" or c["excluded"]: continue           # cost model measures Characters only
    if c["power"] is None or c["level"] is None or c["cost"] is None or c["soul"] is None: continue
    ab = ra(c)
    if not ab: continue
    # Build this card's list of ability signatures and record every occurrence per signature.
    sigs = []
    for i, a in enumerate(ab):
        mk = "".join(a.get("markers") or [])
        sig = mk + " :: " + gen(a.get("text", ""))       # signature = markers + generalized text
        sigs.append(sig)
        variant_cards[sig].append(c["card_number"])
        variant_occ[sig].append((c["card_number"], i))
        variant_text.setdefault(sig, (mk, gen(a.get("text", ""))))
    delta = pb(c) - c["power"]                            # total power this card spent across its abilities
    e = era.get(c["card_number"])
    cards.append((c["card_number"], delta, sigs, e))
    # A SINGLE-ability card measures its one ability's cost DIRECTLY (delta == that ability's cost).
    # These "isolated" deltas seed STEP 1. (NOTE: the modern bucket "m" never fills anymore — see the
    # superseded header: card_era.json no longer emits "modern".)
    if len(ab) == 1:
        (iso[sigs[0]]["m"] if e == "modern" else iso[sigs[0]]["l"]).append(delta)

ALLV = set(variant_cards)
print(f"total variants (distinct abilities): {len(ALLV)}")
print(f"valid cards: {len(cards)}  | abilities with reliable official EN: {len(EN_AB)}")

# ---------- STEP 1: MEASURED (isolated) ----------
# The most reliable costs: taken directly from single-ability cards. cost = the MODE (rounded to 500)
# of those isolated deltas. (Legacy of the era split: prefer >=2 modern samples, else pool all; with
# no modern samples anymore, `use` is always the full pool.)
cost = {}; method = {}; nsamp = {}; rng = {}; eralab = {}
for sig, d in iso.items():
    use = d["m"] if len(d["m"]) >= 2 else (d["m"] + d["l"])
    if not use: continue
    cost[sig] = mode500(use); method[sig] = "measured"; nsamp[sig] = len(use)
    rng[sig] = (min(use), max(use)); eralab[sig] = "modern" if len(d["m"]) >= 2 else ("legacy" if not d["m"] else "mix")
print(f"STEP1 measured: {len(cost)} variants")

# ---------- STEP 2: RESIDUAL with propagation ----------
# Cost abilities that never appear alone, by subtraction. On a multi-ability card where only ONE
# ability is still unknown, that unknown's cost = card delta - sum(known ability costs). Each newly
# solved ability becomes "known", so we iterate (up to 10 passes) until no further ability resolves.
multi = [(cn, dl, sg, e) for (cn, dl, sg, e) in cards if len(sg) > 1]
for it in range(10):
    res = collections.defaultdict(list)
    for cn, dl, sg, e in multi:
        unknown = [s for s in sg if s not in cost]
        if len(unknown) == 1:                                 # exactly one unknown -> solvable this pass
            known_sum = sum(cost[s] for s in sg if s in cost)
            res[unknown[0]].append(dl - known_sum)            # one residual sample for that ability
    new = 0
    for sig, samples in res.items():
        if sig in cost: continue
        cost[sig] = mode500(samples); method[sig] = "residual"; nsamp[sig] = len(samples)
        rng[sig] = (min(samples), max(samples)); eralab[sig] = "—"; new += 1
    print(f"  STEP2 iter {it+1}: +{new} variants via residual (total {len(cost)})")
    if new == 0: break                                        # fixed point reached

# ---------- VALIDATION: error on fully-known multi cards ----------
# Sanity check the additive model: for every multi card whose abilities are ALL priced, does the sum
# reconstruct the real delta? Report the share within one 500-step (the accuracy claim).
errs = []
for cn, dl, sg, e in multi:
    if all(s in cost for s in sg):
        errs.append(abs(dl - sum(cost[s] for s in sg)))
if errs:
    within = sum(1 for x in errs if x <= 500) / len(errs)
    print(f"VALIDATION (known multi cards n={len(errs)}): |error|<=500 in {within*100:.0f}% | mean err {sum(errs)/len(errs):.0f}")

# ---------- STEP 3: ESTIMATED (family median) ----------
# Last resort for abilities never measured or resolved: use the median cost of their family as a
# stand-in (LOW confidence). Family medians are computed from the measured+residual costs above.
fam_known = collections.defaultdict(list)
for sig, cst in cost.items():
    fam_known[family(variant_text[sig][1])].append(cst)
import statistics as st
fam_med = {f: r500(st.median(v)) for f, v in fam_known.items()}
for sig in ALLV:
    if sig in cost: continue
    f = family(variant_text[sig][1])
    cost[sig] = fam_med.get(f, 500); method[sig] = "estimated"; nsamp[sig] = 0   # 500 fallback if family unseen
    rng[sig] = (None, None); eralab[sig] = "—"
print(f"STEP3 estimated: {sum(1 for s in ALLV if method[s]=='estimated')} variants (rest)")

# ---------- EN per variant: any occurrence (card,idx) with reliable EN ----------
# One EN string per signature: reuse the official EN of the FIRST occurrence that has a verified match
# (all occurrences share the same generalized JP, so any verified EN represents the whole variant).
def variant_en(sig):
    for occ in variant_occ[sig]:
        if occ in EN_AB: return EN_AB[occ]
    return ""
def variant_ex(sig):   # example card to show: prefer one with EN, else the first card that has the ability
    for cn, idx in variant_occ[sig]:
        if (cn, idx) in EN_AB: return cn
    return variant_cards[sig][0]

# ---------- build rows ----------
# Confidence downgrades from the method: MEASURED can reach HIGH (many tight samples); RESIDUAL caps at
# MEDIUM (derived, not observed); ESTIMATED is always LOW. Tight range (<=1000 spread) + enough samples
# is what promotes a tier. (This is the OLD confidence rule; the canonical model in cost_model.py uses
# a different evidence-based definition — see COST_MODEL.md.)
def conf(sig):
    m = method[sig]
    if m == "measured":
        lo, hi = rng[sig]
        if nsamp[sig] >= 5 and hi - lo <= 1000: return "HIGH"     # well-sampled and consistent
        if nsamp[sig] >= 2: return "MEDIUM"
        return "LOW"
    if m == "residual":
        lo, hi = rng[sig]
        if nsamp[sig] >= 3 and hi - lo <= 1000: return "MEDIUM"
        return "LOW"
    return "LOW"

rows = []
for sig in ALLV:
    mk, txt = variant_text[sig]
    lo, hi = rng[sig]
    rows.append({"fam": family(txt), "type": mk, "jp": txt, "en": variant_en(sig),
                 "cost": cost[sig], "method": method[sig], "n": nsamp[sig],
                 "range": (f"{lo}..{hi}" if lo is not None else "—"), "era": eralab[sig],
                 "ex": variant_ex(sig), "conf": conf(sig), "ncards": len(variant_cards[sig])})
order = {"measured": 0, "residual": 1, "estimated": 2}
rows.sort(key=lambda r: (r["fam"], order[r["method"]], -r["cost"], -r["n"]))

# ---------- Excel ----------
wb = Workbook(); ws = wb.active; ws.title = "All abilities"
hdr = ["Family", "Type", "Ability (JP real, generalized)", "Official EN",
       "Cost (500s)", "Method", "Confidence", "n", "Range", "Era", "# cards", "Example card"]
ws.append(hdr)
for c in range(1, len(hdr)+1):
    cell = ws.cell(1, c); cell.fill = PatternFill("solid", fgColor="2F5597")
    cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(wrap_text=True, vertical="top")
cf = {"HIGH": "C6EFCE", "MEDIUM": "FFEB9C", "LOW": "FFC7CE"}
mf = {"measured": "D9E1F2", "residual": "FCE4D6", "estimated": "F2F2F2"}
for r in rows:
    ws.append([r["fam"], r["type"], r["jp"], r["en"], r["cost"], r["method"], r["conf"],
               r["n"], r["range"], r["era"], r["ncards"], r["ex"]])
    ws.cell(ws.max_row, 5).font = Font(bold=True)
    ws.cell(ws.max_row, 6).fill = PatternFill("solid", fgColor=mf[r["method"]])
    ws.cell(ws.max_row, 7).fill = PatternFill("solid", fgColor=cf[r["conf"]])
for col, w in zip("ABCDEFGHIJKL", [15, 10, 60, 58, 11, 10, 10, 5, 12, 8, 8, 13]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# summary sheet
ws2 = wb.create_sheet("Summary")
tot = len(rows)
bym = collections.Counter(r["method"] for r in rows)
byc = collections.Counter(r["conf"] for r in rows)
byf = collections.Counter(r["fam"] for r in rows)
ws2.append(["MASTER ABILITY LIST — Weiss Schwarz (custom cards)"]); ws2["A1"].font = Font(bold=True, size=13)
for line in ["", f"Total abilities (variants): {tot}", "",
    "By method:", f"  measured (direct delta, single-ability cards): {bym['measured']}",
    f"  residual (subtract known seeds, propagated): {bym['residual']}",
    f"  estimated (family median): {bym['estimated']}", "",
    "By confidence:", f"  HIGH: {byc['HIGH']}   MEDIUM: {byc['MEDIUM']}   LOW: {byc['LOW']}", "",
    f"With verified official EN: {sum(1 for r in rows if r['en'])}", "",
    "Read the 'How to use' sheet to interpret cost, method and confidence."]:
    ws2.append([line])
ws2.append([""]); ws2.append(["By family (top):"])
for f, n in byf.most_common():
    ws2.append([f"  {f}", n])
ws2.column_dimensions["A"].width = 52

ws3 = wb.create_sheet("How to use")
for line in ["HOW TO USE THIS LIST", "",
    "What it is: balance reference for CUSTOM cards. 'I want this effect -> it costs X power'.",
    "Cost = power SUBTRACTED from the card (power_real = power_base - cost). Always a multiple of 500.",
    "power_base = 3000 + 2500·level + 1500·cost − 1000·(soul trigger) − 1000·(soul−1).", "",
    "COLUMNS:",
    "  MEASURED method  = cost measured directly on SINGLE-ability cards (most reliable).",
    "  RESIDUAL method= the ability only appears alongside others; the known ones are subtracted and its cost remains.",
    "  ESTIMATED method= not yet measurable; the family median is used (indicative, LOW confidence).",
    "  HIGH/MEDIUM/LOW confidence according to number of samples and range dispersion.",
    "  n = samples used; Range = min..max of the measurements; # cards = in how many cards it appears.", "",
    "PRINCIPLES (from the whole session):",
    "  • Resource economy: card to hand/stock = +1 resource ≈ +1000; to waiting = you lose a resource.",
    "  • Era: legacy (<2017) ≈ 2× the modern cost (powercreep). Design with MODERN values.",
    "  • Bundle = SUM; modal 'choose 1 of N' = the strongest option; multi-trigger = value × number of triggers.",
    "  • CX-combo / hard-gate: floor ~500 regardless of power (paid by assembling the combo).",
    "  • For NOVEL effects: decompose into primitives, apply modifiers, compose, round to 500.",
    "    (see Ability_Cost_Guide.xlsx for the full model and examples)"]:
    ws3.append([line])
ws3["A1"].font = Font(bold=True, size=12); ws3.column_dimensions["A"].width = 100

out = os.path.join(D, "Complete_Abilities_List.xlsx"); wb.save(out)
print(f"\nTOTAL rows: {tot} | measured {bym['measured']} residual {bym['residual']} estimated {bym['estimated']}")
print(f"Other (family): {byf['Other']} | with EN: {sum(1 for r in rows if r['en'])}")
print("written:", out)
